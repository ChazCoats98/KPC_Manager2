import sys
import numpy as np
import shutil
import hashlib
import re
from PyPDF2 import PdfReader
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer
from datetime import datetime, timedelta, date
import time
import typing
from pymongo import MongoClient
from PyQt5.QtWidgets import QGridLayout, QHBoxLayout, QVBoxLayout, QFormLayout, QPushButton, QWidget, QLineEdit, QLabel, QTableWidget, QTableWidgetItem, QDockWidget, QHeaderView, QFileSystemModel, QComboBox
from PyQt5 import QtGui
from PyQt5.QtCore import QModelIndex, Qt, QDir, QAbstractItemModel, Qt, pyqtSignal, QSortFilterProxyModel, QDate
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QMessageBox,
    QTableView,
    QTreeView,
    QTreeWidget,
    QTreeWidgetItem,
    QFileDialog,
    QCheckBox,
    QScrollArea
)
from utils import database
from openpyxl import load_workbook

emailRegex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
passwordRegex = r'^(?=.*?[A-Z])(?=.*?[a-z])(?=.*?[0-9])(?=.*?[#?!@$%^&*-]).{8,}$'

class loginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("KPC Manager")
        self.resize(400, 200)
        
        layout = QGridLayout()
        
        # Email form 
        emailLabel = QLabel('Email:')
        self.emailInput = QLineEdit()
        self.emailInput.setPlaceholderText('Enter Email')
        layout.addWidget(emailLabel, 0, 0)
        layout.addWidget(self.emailInput, 0, 1)
        
        # Password form
        passwordLabel = QLabel('Password:')
        self.passwordInput = QLineEdit()
        self.passwordInput.setPlaceholderText('Enter Password')
        layout.addWidget(passwordLabel, 1, 0)
        layout.addWidget(self.passwordInput, 1, 1)
        
        # Login Button
        loginButton = QPushButton('Login')
        loginButton.clicked.connect(self.loginSubmit)
        layout.addWidget(loginButton, 2, 0, 1, 2)
        
        registerButton = QPushButton('Register')
        registerButton.clicked.connect(self.registerSubmit)
        layout.addWidget(registerButton, 3, 0, 1, 2)
        
        self.setLayout(layout)
        
        # function to handle button submit
    def loginSubmit(self):
        email = self.emailInput.text()
        pwd = self.passwordInput.text()
            
        login(email, pwd)
        
    def registerSubmit(self):
        global r
        r = registerWindow()
        r.show()
        window.close()
        
class registerWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("KPC Manager")
        self.resize(400, 200)
        
        layout = QGridLayout()
        
        # Email form 
        emailLabel = QLabel('Email:')
        self.emailInput = QLineEdit()
        self.emailInput.setPlaceholderText('Enter Email')
        layout.addWidget(emailLabel, 0, 0)
        layout.addWidget(self.emailInput, 0, 1)
        
        # Password form
        passwordLabel = QLabel('Password:')
        self.passwordInput = QLineEdit()
        self.passwordInput.setPlaceholderText('Enter Password')
        layout.addWidget(passwordLabel, 1, 0)
        layout.addWidget(self.passwordInput, 1, 1)
        
        # Repeat Password form
        repeatPasswordLabel = QLabel('Repeat Password:')
        self.repeatPasswordInput = QLineEdit()
        self.repeatPasswordInput.setPlaceholderText('repeat Password')
        layout.addWidget(repeatPasswordLabel, 2, 0)
        layout.addWidget(self.repeatPasswordInput, 2, 1)
        
        # Register Button
        registerButton = QPushButton('Register')
        registerButton.clicked.connect(self.registerUser)
        layout.addWidget(registerButton, 3, 0, 1, 2)
        
        self.setLayout(layout)
        
        # function to handle button submit
    def registerUser(self):
        email = self.emailInput.text()
        pwd = self.passwordInput.text()
        confPwd = self.repeatPasswordInput.text()
            
        register(email, pwd, confPwd)
        
class partForm(QWidget):
    partSubmitted = pyqtSignal()
    def __init__(self, mode="add", partId=None):
        super().__init__()
        self.mode = mode
        self.partId = partId
        self.setWindowTitle("Part")
        self.resize(800, 600)
        
        layout = QGridLayout()
        
        # Part number form 
        partLabel = QLabel('Part Number:')
        self.partInput = QLineEdit()
        self.partInput.setPlaceholderText('Enter part number')
        layout.addWidget(partLabel, 0, 0)
        layout.addWidget(self.partInput, 1, 0)
        
        # Part revision form
        revLabel = QLabel('Revision Letter:')
        self.revInput = QLineEdit()
        self.revInput.setPlaceholderText('Enter revision letter')
        layout.addWidget(revLabel, 0, 1)
        layout.addWidget(self.revInput, 1, 1)
        
        # upload date form
        udLabel = QLabel('Last Net-Inspect Upload Date:')
        self.udInput = QLineEdit()
        self.udInput.setPlaceholderText('Enter upload date')
        layout.addWidget(udLabel, 0, 2)
        layout.addWidget(self.udInput, 1, 2)
        
        # Notes form
        notesLabel = QLabel('Notes:')
        self.notesInput = QLineEdit()
        self.notesInput.setPlaceholderText('Enter notes')
        layout.addWidget(notesLabel, 0, 3)
        layout.addWidget(self.notesInput, 1, 3)
        
        #No current manufacturing flag
        self.manufacturingCheck = QCheckBox(text="No Current Manufacturing")
        layout.addWidget(self.manufacturingCheck, 1, 4)
        
        # Submit button Button
        addFeatureButton = QPushButton('Add Feature')
        addFeatureButton.clicked.connect(self.addFeature)
        layout.addWidget(addFeatureButton, 5, 1, 1, 3)
        
        addPartButton = QPushButton('Save Part')
        addPartButton.setStyleSheet("background-color: #3ADC73")
        addPartButton.clicked.connect(self.submitPart)
        layout.addWidget(addPartButton, 6, 0, 1, 5)
        
        cancelButton = QPushButton('Cancel')
        cancelButton.setStyleSheet("background-color: #D6575D")
        cancelButton.clicked.connect(self.closeWindow)
        layout.addWidget(cancelButton, 7, 0, 1, 5)
        
        self.featureTable = QTableWidget()
        self.featureTable.setColumnCount(6)
        self.featureTable.setHorizontalHeaderLabels(["Feature Number", "KPC Designation", "KPC Number", "Operation Number", "Tolerance", "Engine"])
        self.featureTable.horizontalHeader().setStretchLastSection(False)
        for column in range(self.featureTable.columnCount()):
            self.featureTable.horizontalHeader().setSectionResizeMode(column, QHeaderView.Stretch)
            
        layout.addWidget(self.featureTable, 4, 0, 1, 5)
        
        
        self.setLayout(layout)
    def addFeature(self):
            self.featureForm = FeatureForm(self)
            self.featureForm.show()
            
    def addFeatureToTable(self, feature_data):
        row_position = self.featureTable.rowCount()
        self.featureTable.insertRow(row_position)
        
        keys = ['feature', 'designation', 'kpcNum', 'opNum', 'tol', 'engine']
        for i, key in enumerate(keys):
            value = feature_data.get(key, '')
            self.featureTable.setItem(row_position, i, QTableWidgetItem(value))
            
    def submitPart(self):
                
        upload_date_str = self.udInput.text().strip()
        if not upload_date_str:
            QMessageBox.warning(self, "Error", "Upload date cannot be empty.")
            return
        try: 
            upload_date = datetime.strptime(upload_date_str, '%m/%d/%Y')
        except ValueError:
            QMessageBox.warning(self, "Error", "Upload date must be in MM/DD/YYYY Format.")
            return
            
        due_date = upload_date + timedelta(days=90)
        due_date_str = due_date.strftime('%m/%d/%Y')
        new_part_data = {
            "partNumber": self.partInput.text(),
            "rev": self.revInput.text(),
            "uploadDate": self.udInput.text(),
            "dueDate": due_date_str,
            "notes": self.notesInput.text(),
            "currentManufacturing": self.manufacturingCheck.isChecked(),
            "features": []
        }
        for row in range(self.featureTable.rowCount()):
            feature = {
                "feature": self.featureTable.item(row, 0).text(),
                "designation": self.featureTable.item(row, 1).text(),
                "kpcNum": self.featureTable.item(row, 2).text(),
                "opNum": self.featureTable.item(row, 3).text(),
                "tol": self.featureTable.item(row, 4).text(),
                "engine": self.featureTable.item(row, 5).text(),
            }
            new_part_data["features"].append(feature)
            
        def on_submit_success(is_success):
            if is_success:
                self.partSubmitted.emit()
        if self.mode == "add":
            if database.check_for_part(self.partInput.text()):
                QMessageBox.warning(self, "Error", "Part already exists in database.")
                return
            else: 
                database.submit_new_part(new_part_data, callback=on_submit_success)
        elif self.mode == "edit":
            database.update_part_by_id(self.partId, new_part_data, callback=on_submit_success)
            
        self.close()
    
    def loadPartData(self, selectedPartData):
        self.partInput.setText(selectedPartData['partNumber'])
        self.revInput.setText(selectedPartData['rev'])
        self.udInput.setText(selectedPartData['uploadDate'])
        self.notesInput.setText(selectedPartData['notes'])
        self.manufacturingCheck.setChecked(selectedPartData.get('currentManufacturing', False))
        
        self.featureTable.setRowCount(0)
        for feature in selectedPartData['features']:
            self.addFeatureToTable(feature)
    
    def closeWindow(self):
        self.close()
        
class dashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("KPC Manager")
        self.resize(1200, 800)
        
        self.mainWidget = QWidget(self)
        self.mainLayout = QVBoxLayout()
        
        self.tree_view = PartTreeView(self)
        self.part_data = database.get_all_data()
        self.model = PartFeaturesModel(self.part_data)
        self.proxyModel = QSortFilterProxyModel(self)
        self.proxyModel.setSourceModel(self.model)
        self.tree_view.setModel(self.proxyModel)
        self.tree_view.setSortingEnabled(True)
        self.tree_view.sortByColumn(3, Qt.AscendingOrder)
        self.tree_view.resize(1200,800)
        self.mainLayout.addWidget(self.tree_view)
        
        self.addPart = QPushButton('Add Part')
        self.addPart.setStyleSheet("background-color: #3ADC73")
        self.addPart.clicked.connect(self.openPartForm)
        self.mainLayout.addWidget(self.addPart)
        
        self.editPart = QPushButton('Edit Selected Part')
        self.editPart.setStyleSheet("background-color: #DFDA41")
        self.editPart.clicked.connect(self.editSelectedPart)
        self.mainLayout.addWidget(self.editPart)
        
        self.uploadPartData = QPushButton('Upload Data for selected Part')
        self.uploadPartData.setStyleSheet("background-color: #E6A42B")
        self.uploadPartData.clicked.connect(self.openUploadForm)
        self.mainLayout.addWidget(self.uploadPartData)
        
        self.showHistoricalUploads = QPushButton('Show Past Data Uploads for selected Part')
        self.showHistoricalUploads.setStyleSheet("background-color: #439EF3")
        self.showHistoricalUploads.clicked.connect(self.openHistoricalUploadWindow)
        self.mainLayout.addWidget(self.showHistoricalUploads)
        
        self.deletePart = QPushButton('Delete Part')
        self.deletePart.setStyleSheet("background-color: #D6575D")
        self.deletePart.clicked.connect(self.deleteSelectedPart)
        self.mainLayout.addWidget(self.deletePart)
        
        self.mainWidget.setLayout(self.mainLayout)
        self.setCentralWidget(self.mainWidget)
        
    def openPartForm(self):
        self.partForm = partForm()
        self.partForm.partSubmitted.connect(self.refreshTreeView)
        self.partForm.show()
        
    def openUploadForm(self):
        index = self.tree_view.currentIndex()
        if not index.isValid():
            QMessageBox.warning(self, "Selection", "No part selected.")
            return
        sourceIndex = self.proxyModel.mapToSource(index)
        part_id = self.model.getPartId(sourceIndex)
        if part_id is None:
            QMessageBox.warning(self, "Error", "Failed to identify selected part.")
            return
        
        selectedPartData = database.get_part_by_id(part_id)
            
        if not selectedPartData:
            QMessageBox.warning(self, "Error", "Could not find part data.")
            return
        
        self.uploadForm = uploadDataForm(partId = part_id)
        self.uploadForm.loadPartData(selectedPartData)
        self.uploadForm.dataSubmitted.connect(self.refreshTreeView)
        self.uploadForm.show()
        
    def openHistoricalUploadWindow(self):
        index = self.tree_view.currentIndex()
        if not index.isValid():
            QMessageBox.warning(self, "Error", "No part selected.")
            return
        sourceIndex = self.proxyModel.mapToSource(index)
        part_id = self.model.getPartId(sourceIndex)
        if part_id is None:
            QMessageBox.warning(self, "Error", "Failed to identify selected part.")
            return
        
        selectedPartData = database.get_part_by_id(part_id)
        selectedPartUploadData = database.get_measurements_by_id(part_id)
        print(selectedPartUploadData)
        if not selectedPartData or selectedPartUploadData is None:
            QMessageBox.warning(self, "Error", "Could not find part data.")
            return
        
        try: 
            self.historicalData = historicalData(partId=part_id)
            self.historicalData.loadPartData(selectedPartData, selectedPartUploadData)
            database.delete_duplicate_measurements()
            self.historicalData.show()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            print(e)
        
    def deleteSelectedPart(self):
        index = self.tree_view.currentIndex()
        if not index.isValid():
            QMessageBox.warning(self, "Selection", "No part selected.")
            return
        sourceIndex = self.proxyModel.mapToSource(index)
        part_id = self.model.getPartId(sourceIndex)
        if part_id is None:
            QMessageBox.warning(self, "Error", "Failed to identify selected part.")
            return
        database.delete_part(self, part_id)
        
    def editSelectedPart(self):
        index = self.tree_view.currentIndex()
        if not index.isValid():
            QMessageBox.warning(self, "Selection", "No part selected.")
            return
        sourceIndex = self.proxyModel.mapToSource(index)
        part_id = self.model.getPartId(sourceIndex)
        if part_id is None:
            QMessageBox.warning(self, "Error", "Failed to identify selected part.")
            return
        
        selectedPartData = database.get_part_by_id(part_id)
        if not selectedPartData:
            QMessageBox.warning(self, "Error", "Could not find part data.")
            return
        
        self.partForm = partForm(mode="edit", partId=part_id)
        self.partForm.loadPartData(selectedPartData)
        self.partForm.partSubmitted.connect(self.refreshTreeView)
        self.partForm.show()       
        
        
    def refreshTreeView(self):
        updated_parts_data = database.get_all_data()
        
        self.model.updateData(updated_parts_data)
        
class PartFeaturesModel(QAbstractItemModel):
    def __init__(self, part_data, parent=None):
        super(PartFeaturesModel, self).__init__(parent)
        
        self.part_data = part_data if part_data is not None else []
        
    def index(self, row, column, parent=QModelIndex()):
        if not self.hasIndex(row, column, parent):
            return QModelIndex()
        
        if not parent.isValid():
            part = self.part_data[row]
            return self.createIndex(row, column, part)
        else: 
            part = parent.internalPointer()
            feature = part['features'][row]
            return self.createIndex(row, column, feature)
        
    def parent(self, index):
        if not index.isValid():
            return QModelIndex()
        
        child = index.internalPointer()
        if  'features' in child:
            
            return QModelIndex()
        else:
            for part in self.part_data:
                if child in part.get('features', []):
                    parent_row = self.part_data.index(part)
                    return self.createIndex(parent_row, 0, part)
        return QModelIndex()
    
    def rowCount(self, parent=QModelIndex()):
        if parent.column() > 0:
            return 0
        if not parent.isValid():
            return len(self.part_data)
        else: 
            part = parent.internalPointer()
            return len(part.get('features', []))
        
    def columnCount(self, parent=QModelIndex()):
        return 6
    
    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        item = index.internalPointer()
        if role == Qt.DisplayRole:
            if 'feature' in item:
                if index.column() == 0:
                    return f"Feature #: {item['feature']}"
                elif index.column() == 1:
                    return f"KPC Designation: {item['designation']}"
                elif index.column() == 2:
                    return f"KPC Number: {item['kpcNum']}"
                elif index.column() == 3:
                    return f"Tolerance Value: {item['tol']}"
                elif index.column() == 4:
                    return f"Engine: {item['engine']}"
                elif index.column() == 5:
                    return f"CPK: {item.get('cpk', 'N/A')}"
            else:
                if index.column() == 0:
                    return item.get('partNumber', '')
                elif index.column() == 1:
                    return item.get('rev', '')
                elif index.column() == 2:
                    date_string = item.get('uploadDate')
                    if date_string:
                        formatted_date = format_date(date_string)
                    return formatted_date
                elif index.column() == 3:
                    return item.get('dueDate', '')
                elif index.column() == 4:
                    return item.get('notes', '')
            pass
        elif role == Qt.BackgroundRole:
            if 'currentManufacturing' in item and item['currentManufacturing']:
                return QtGui.QBrush(QtGui.QColor('red'))     
        return None
    
    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            headers = ['Part Number', 'Revision', 'Last Upload Date', 'Upload Due Date', 'notes']
            if section < len(headers):
                return headers[section]
        return None
            
    def getPartId(self, index):
        if not index.isValid():
            return None
        item = index.internalPointer()
        
        return item.get('partNumber')
    
    def updateData(self, new_data):
        self.beginResetModel()
        self.part_data = new_data
        self.endResetModel()
        
        
class PartTreeView(QTreeView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.header().setStretchLastSection(False)
        self.header().setSectionResizeMode(QHeaderView.Stretch)
        self.setUniformRowHeights(True)
            
    def resizeEvent(self, event):
        super().resizeEvent(event)
        
    def setModel(self, model):
        super().setModel(model)
        
class FeatureForm(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.setWindowTitle("Add Feature")
        self.resize(600, 100)
        
        layout = QGridLayout()
        
        # Part number form 
        featureLabel = QLabel('Feature Number:')
        self.featureInput = QLineEdit()
        self.featureInput.setPlaceholderText('Enter Feature Number')
        layout.addWidget(featureLabel, 0, 0)
        layout.addWidget(self.featureInput, 1, 0)
        
        # Part revision form
        kpcLabel = QLabel('KPC Designation:')
        self.kpcInput = QLineEdit()
        self.kpcInput.setPlaceholderText('Enter KPC Designation')
        layout.addWidget(kpcLabel, 0, 1)
        layout.addWidget(self.kpcInput, 1, 1)
        
        opNumLabel = QLabel('Op Number:')
        self.opNumInput = QLineEdit()
        self.opNumInput.setPlaceholderText('Enter KPC Designation')
        layout.addWidget(opNumLabel, 0, 2)
        layout.addWidget(self.opNumInput, 1, 2)
        
        # upload date form
        kpcNumLabel = QLabel('KPC Number:')
        self.kpcNumInput = QLineEdit()
        self.kpcNumInput.setPlaceholderText('Enter KPC Number from Net-Inspect')
        layout.addWidget(kpcNumLabel, 0, 3)
        layout.addWidget(self.kpcNumInput, 1, 3)
        
        # Notes form
        requirementLabel = QLabel('Requirement:')
        self.requirementInput = QLineEdit()
        self.requirementInput.setPlaceholderText('Enter Blueprint Requirement')
        layout.addWidget(requirementLabel, 0, 4)
        layout.addWidget(self.requirementInput, 1, 4)
        
        engineLabel = QLabel('Engine:')
        self.engineInput = QLineEdit()
        self.engineInput.setPlaceholderText('Enter Part Engine Program')
        layout.addWidget(engineLabel, 0, 45)
        layout.addWidget(self.engineInput, 1, 5)
        
        # Submit button Button
        addFeatureButton = QPushButton('Add Feature')
        addFeatureButton.setStyleSheet("background-color: #3ADC73")
        addFeatureButton.clicked.connect(self.submitFeature)
        layout.addWidget(addFeatureButton, 2, 2, 1, 1)
        
        cancelButton = QPushButton('Cancel')
        cancelButton.setStyleSheet("background-color: #D6575D")
        cancelButton.clicked.connect(self.closeWindow)
        layout.addWidget(cancelButton, 3, 2, 1, 1)
        
        self.setLayout(layout)
        
    def submitFeature(self):
        
        feature_data = {
            "feature": self.featureInput.text(),
            "designation": self.kpcInput.text(),
            "opNum": self.opNumInput.text(),
            "kpcNum": self.kpcNumInput.text(),
            "tol": self.requirementInput.text(),
            "engine": self.engineInput.text(),
        }
        if self.parent:
            self.parent.addFeatureToTable(feature_data)
            
        self.close()
        
    def closeWindow(self):
        self.close()
        
class uploadDataForm(QWidget):
    dataSubmitted = pyqtSignal()
    def __init__(self, partId=None):
        super().__init__()
        self.partId = partId
        self.serialNumberInputs = []
        self.featureTables = []
        self.setWindowTitle("New Data Upload")
        self.resize(800, 600)
        
        layout = QGridLayout()
        
        self.scrollArea = QScrollArea(self)
        self.scrollAreaWidgetContents = QWidget()
        self.scrollArea.setWidgetResizable(True)
        self.scrollAreaWidgetLayout = QVBoxLayout(self.scrollAreaWidgetContents)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        layout.addWidget(self.scrollArea, 4, 0, 1, 7)
        
        # Part number form 
        partLabel = QLabel('Part Number:')
        self.partNumber = QLabel('')
        layout.addWidget(partLabel, 0, 0)
        layout.addWidget(self.partNumber, 0, 1)
        
        # Part revision form
        revLabel = QLabel('Revision Letter:')
        self.revLetter = QLabel('')
        layout.addWidget(revLabel, 0, 2)
        layout.addWidget(self.revLetter, 0, 3)
        
        # upload date form
        udLabel = QLabel('Last Net-Inspect Upload Date:')
        self.uploadDate = QLabel('')
        layout.addWidget(udLabel, 0, 4)
        layout.addWidget(self.uploadDate, 0, 5)
        
        serialNumberLabel = QLabel('Serial Number:')
        self.serialNumberInput = QLineEdit()
        self.serialNumberInput.setPlaceholderText('Enter Serial Number')
        layout.addWidget(serialNumberLabel, 2, 0)
        layout.addWidget(self.serialNumberInput, 2, 1, 1, 2)
        
        self.serialNumberInput.textChanged.connect(self.checkSerialNumber)
        
        runNumberLabel = QLabel('Run Number:')
        self.runNumberInput = QLineEdit()
        self.runNumberInput.setPlaceholderText('Enter Run Number')
        layout.addWidget(runNumberLabel, 2, 4)
        layout.addWidget(self.runNumberInput, 2, 5, 1, 2)
        
        machineLabel = QLabel('Machine:')
        self.machineComboBox = QComboBox()
        self.machineComboBox.addItems([ '', 'CMM - Zeiss Accura'])
        layout.addWidget(machineLabel, 3, 0)
        layout.addWidget(self.machineComboBox, 3, 1)
        
        lotSizeLabel = QLabel('Lot Size:')
        self.lotSizeComboBox = QComboBox()
        self.lotSizeComboBox.addItems(['', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25'])
        layout.addWidget(lotSizeLabel, 3, 4)
        layout.addWidget(self.lotSizeComboBox, 3, 5, 1, 1)
        
        self.serialNumbersLayout = QVBoxLayout()
        layout.addLayout(self.serialNumbersLayout, 4,0,1,7)
        
        self.lotSizeComboBox.currentTextChanged.connect(self.onLotSizeChange)
        
        readPdfButton = QPushButton('Add Data From PDF')
        readPdfButton.setStyleSheet("background-color: #439EF3")
        readPdfButton.clicked.connect(self.openPdfFileDialog)
        layout.addWidget(readPdfButton, 6, 2, 1, 3)
        
        addPartButton = QPushButton('Save Data')
        addPartButton.setStyleSheet("background-color: #3ADC73")
        addPartButton.clicked.connect(self.submitData)
        layout.addWidget(addPartButton, 7, 2, 1, 3)
        
        cancelButton = QPushButton('Cancel')
        cancelButton.setStyleSheet("background-color: #D6575D")
        cancelButton.clicked.connect(self.closeWindow)
        layout.addWidget(cancelButton, 8, 2, 1, 3)
        
        self.dataTable = QTableWidget()
        self.dataTable.setColumnCount(5)
        self.dataTable.horizontalHeader().setStretchLastSection(False)
        self.dataTable.setHorizontalHeaderLabels(['Feature Number', 'KPC Number', 'Blueprint Dimension', 'Op Number', 'Measurement'])
        for column in range(self.dataTable.columnCount()):
            self.dataTable.horizontalHeader().setSectionResizeMode(column, QHeaderView.Stretch)
            
            
        
        
        self.setLayout(layout)
        
    def onLotSizeChange(self, text):
        if text.isdigit():
            lot_size = int(text)
            self.createLotInputs(lot_size)
        else:
            self.clearLotInputs()
            
    def createLotInputs(self, lot_size):
        self.clearLotInputs()
        
        partData = database.get_part_by_id(self.partId)
        if not partData:
            QMessageBox.warning(self, "Error", "Part data not found in database.")
            return
        part_features = partData['features']
        
        for lot_index in range(lot_size):
            serialNumberInput = QLineEdit()
            serialNumberInput.setPlaceholderText(f'Enter Serial Number {lot_index+1}')
            self.serialNumbersLayout.addWidget(serialNumberInput)
            self.serialNumberInputs.append(serialNumberInput)
            
            featureTable = QTableWidget()
            featureTable.setColumnCount(5)
            featureTable.setRowCount(len(part_features))
            featureTable.setHorizontalHeaderLabels(['Feature Number', 'KPC Number', 'Blueprint Dimension', 'Op Number', 'Measurement'])
            featureTable.horizontalHeader().setStretchLastSection(False)
            for column in range(featureTable.columnCount()):
                featureTable.horizontalHeader().setSectionResizeMode(column, QHeaderView.Stretch)
                
            featureTable.setFixedHeight(250)
                
            for row, feature in enumerate(part_features):
                featureTable.setItem(row, 0, QTableWidgetItem(feature['feature']))
                featureTable.setItem(row, 1, QTableWidgetItem(feature['kpcNum']))
                featureTable.setItem(row, 2, QTableWidgetItem(feature['tol']))
                featureTable.setItem(row, 3, QTableWidgetItem(feature.get('opNum', '')))
                featureTable.setItem(row, 4, QTableWidgetItem(''))
                
            
            self.adjustTableHeight(featureTable)
            self.featureTables.append(featureTable)
            
            self.scrollAreaWidgetLayout.addWidget(serialNumberInput)
            self.scrollAreaWidgetLayout.addWidget(featureTable)
        
    def clearLotInputs(self):
        while self.scrollAreaWidgetLayout.count():
            item = self.scrollAreaWidgetLayout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.serialNumberInputs = []
        self.featureTables = []
                    
    def checkSerialNumber(self, text):
        exists = database.check_serial_number(text)
        print(exists)
        if exists:
            self.serialNumberInput.setStyleSheet("background-color: red")
            QMessageBox.warning(self, "Serial Number Invalid", "Data for this serial number has already been uploaded.")
        else: 
            self.serialNumberInput.setStyleSheet("background-color: white")
            
    def adjustTableHeight(self, table):
        total_height = table.horizontalHeader().height()
        for i in range(table.rowCount()):
            total_height += table.rowHeight(i)
        
        margin = 4
        total_height += margin
        
        table.setFixedHeight(total_height)
            
    def addFeatureToTable(self, feature_data):
        row_position = self.dataTable.rowCount()
        self.dataTable.insertRow(row_position)
        for i, key in enumerate(['feature','kpcNum', 'tol']):
            self.dataTable.setItem(row_position, i, QTableWidgetItem(feature_data[key]))
            
    def openPdfFileDialog(self):
        partNumber = self.partNumber.text()
        serialNumber = self.serialNumberInput.text()
        initialDir = f'//Server/d/Inspection/CMM Files/Printouts/{partNumber}'
        
        if serialNumber:
            initialDir = f'//Server/d/Inspection/CMM Files/Printouts/{partNumber}/{serialNumber}'
            
        filePath, _ = QFileDialog.getOpenFileName(
            self, 
            "Open PDF File", 
            initialDir, 
            "PDF Files (*.pdf)"
        )
        if filePath:
            self.extractDataFromPdf(filePath)
            
    def extractDataFromPdf(self, filePath):
        for page_layout in extract_pages(filePath):
            for element in page_layout:
                if isinstance(element, LTTextContainer):
                    for text_line in element:
                        text = text_line.get_text().strip()
                        measurements = re.findall(r"Diameter_Circle \d+\.\d+\s*\|-?\s*(\d+\.\d+)", text)
                        print(measurements)
        #part_data = database.get_part_by_id(self.partNumber.text())
        #tolerances = {feature['kpcNum']: parse_tolerance(feature['tol']) for feature in part_data['features']}
        #print(tolerances)
        #try:
            #reader = PdfReader(filePath)
            #text = ''
            #for page in reader.pages:
                #text += page.extract_text() + '\n'
            
            #pattern = re.compile(r'\b\d*\.?\d+\b')
            #measurements = [float(m) for m in pattern.findall(text)]
            
            #matched_measurements = {kpc: [] for kpc in tolerances.keys()}
            #for measurement in measurements:
                #for kpc, tolerance in tolerances.items():
                    #lower, upper = tolerance
                    #if lower is not None and upper is not None and lower < measurement < upper:
                       # matched_measurements[kpc].append(measurement)
                    
            #print(matched_measurements)
            #for row in range(self.dataTable.rowCount()):
                #kpcNum = self.dataTable.item(row, 1).text()
                #if kpcNum in matched_measurements and matched_measurements[kpcNum]:
                    #self.dataTable.setItem(row, 4, QTableWidgetItem(str(matched_measurements[kpcNum][0])))
        #except Exception as e:
            #print(str(e))
            #QMessageBox.critical(self, "Error", f"Failed to read PDF: {str(e)}")
        
        
            
    def submitData(self):
        part_number = self.partNumber.text()
        existing_part_data = database.get_part_by_id(part_number)
        machine = self.machineComboBox.currentText()
        run_number = self.runNumberInput.text()
        lot_size = self.lotSizeComboBox.currentText()
        target_row = 2
        template_path = './utils/Templates/Measurement_Import_template.xlsx'
        upload_date_value = datetime.strftime(date.today(), '%m/%d/%Y')
        upload_date_file_path = datetime.strftime(date.today(), '%m-%d-%Y')
        upload_date = datetime.strptime(upload_date_value, '%m/%d/%Y')
        new_file_path = f'./Results/{part_number}_data_upload_{upload_date_file_path}.xlsx'
        due_date = upload_date + timedelta(days=90)
        due_date_str = due_date.strftime('%m/%d/%Y')

        if not existing_part_data:
            QMessageBox.warning(self, "Error", "Part data not found in database.")
            return
        
        shutil.copy(template_path,  new_file_path)
        
        workbook = load_workbook(filename=new_file_path)
        sheet = workbook.active
        
        existing_features = existing_part_data.get('features', [])
        features_dict = {feature['kpcNum']: feature for feature in existing_features}
        upload_data = {
                    "partNumber": self.partNumber.text(),
                    "serialNumber": self.serialNumberInput.text(),
                    "uploadDate": upload_date_value,
                    "measurements": []
                }
        
        for serial_input, feature_table in zip(self.serialNumberInputs, self.featureTables):
            serial_number = serial_input.text()
            if database.check_serial_number(serial_number):
                QMessageBox.warning(self, "Duplicate Serial Number", f"Serial Number {serial_number} already in database")
                continue
            
            for row in range(feature_table.rowCount()):
                feature_number = feature_table.item(row, 0).text()
                kpcNum = feature_table.item(row, 1).text()
                opNum = feature_table.item(row, 3).text()
                measurement = feature_table.item(row, 4).text()
            
                if part_number:
                    sheet.cell(row=target_row, column=1).value = part_number
                if feature_number:
                    sheet.cell(row=target_row, column=2).value = feature_number
                if machine:
                    sheet.cell(row=target_row, column=3).value = machine
                if run_number:
                    sheet.cell(row=target_row, column=4).value = run_number
                if lot_size:
                    sheet.cell(row=target_row, column=5).value = lot_size
                if measurement:
                    sheet.cell(row=target_row, column=6).value = measurement
                if serial_number:
                    sheet.cell(row=target_row, column=8).value = serial_number
                target_row += 1
                
                updated_part_data = {
                        "uploadDate": upload_date_value,
                        "dueDate": due_date_str,
                        "features": list(features_dict.values())
                    }
                
                updated_part_data['features'].append({
                    'kpcNum': kpcNum,
                    'opNum': opNum
                })
            
                upload_data["measurements"].append({
                    "kpcNum": kpcNum,
                    "measurement": measurement
                })
                    
                def on_submit_success(is_success):
                    if is_success:
                        self.dataSubmitted.emit()
            
                database.update_part_by_id(self.partId, updated_part_data, callback=on_submit_success)
                database.add_measurement(upload_data)
            
        
        workbook.save(filename=new_file_path)
        try: 
            partNumber = self.partNumber.text()
            basePath = f'//server/D/Quality Control/UPPAP Records/Process Cert + Data Collection/Data points/{partNumber}'
            fileName, ok = QFileDialog.getSaveFileName(
                self,
                "Save Excel File",
                basePath,
                "Excel files (*.xlsx)")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An unexpected error occurred: {e}")
        
        if not fileName:
            QMessageBox.information(self, "Save Canceled", "The save operation was cancelled")
            return
        
        if fileName:
            if not fileName.endswith('.xlsx'):
                fileName += '.xlsx'
            try: 
                shutil.move(new_file_path, fileName)
        
            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred while saving the file: {e}")
                
        
            
        self.close()
        self.calculateAndUpdateCpk(self.partId)
    
    def loadPartData(self, selectedPartData):
        self.partNumber.setText(selectedPartData['partNumber'])
        self.revLetter.setText(selectedPartData['rev'])
        self.uploadDate.setText(selectedPartData['uploadDate'])
        
        self.dataTable.setRowCount(0)
        self.dataTable.setHorizontalHeaderLabels(['Feature Number', 'KPC Number', 'Blueprint Dimension', 'Op Number', 'Measurement'])
            
    def calculateAndUpdateCpk(self, partId):
        part_data = database.get_part_by_id(partId)
        if part_data:
            tolerances = {feature['kpcNum']: parse_tolerance(feature.get('tol', '0-0')) for feature in part_data.get('features', [])}
            
        measurement_data = database.get_measurements_by_id(partId)
        
        measurements_by_kpc = {kpc: [] for kpc in tolerances.keys()}
        
        if measurement_data:
            for entry in measurement_data:
                for measurement in entry.get('measurements', []):
                    kpcNum = measurement.get('kpcNum')
                    if kpcNum and kpcNum in measurements_by_kpc:
                        measurements_by_kpc[kpcNum].append(float(measurement['measurement']))
        print(measurements_by_kpc)            
        cpk_values = {}
        for kpc, data in measurements_by_kpc.items():
            usl, lsl = tolerances[kpc]
            print(f'usl: {usl} lsl: {lsl}')
            if usl is not None and lsl is not None:
                cpk = calculate_cpk(data, usl, lsl)
                if cpk is not None:
                    cpk_values[kpc] = cpk
        print(cpk_values)
                
        if cpk_values:
            formatted_cpk_values = {kpc: round(abs(cpk if cpk is not None else 0), 3) for kpc, cpk in cpk_values.items()}
            database.save_cpk_values(partId, formatted_cpk_values)
        
    def closeWindow(self):
        self.close()
        
class historicalData(QWidget):
    def __init__(self, partId=None):
        super().__init__()
        self.partId = partId
        self.setWindowTitle("data Upload History")
        self.resize(800, 600)
        
        layout = QGridLayout()
        
        partLabel = QLabel('Part Number:')
        self.partNumber = QLabel('')
        layout.addWidget(partLabel, 0, 0)
        layout.addWidget(self.partNumber, 0, 1)

        revLabel = QLabel('Revision Letter:')
        self.revLetter = QLabel('')
        layout.addWidget(revLabel, 0, 2)
        layout.addWidget(self.revLetter, 0, 3)
        
        deleteMeasButton = QPushButton('Delete Current Measurement Record')
        deleteMeasButton.setStyleSheet("background-color: #D6575D")
        deleteMeasButton.clicked.connect(self.deleteMeasurement)
        layout.addWidget(deleteMeasButton, 6, 0, 1, 6)
        
        cancelButton = QPushButton('Close Window')
        cancelButton.setStyleSheet("background-color: #D6575D")
        cancelButton.clicked.connect(self.closeWindow)
        layout.addWidget(cancelButton, 7, 0, 1, 6)
        
        self.model = QStandardItemModel()
        self.proxyModel = DateSortProxyModel(dateColumnIndex=0, parent=self)
        self.proxyModel.setSourceModel(self.model)
        self.treeView= QTreeView()
        self.treeView.header().setStretchLastSection(False)
        self.treeView.header().setSectionResizeMode(QHeaderView.Stretch)
        self.model.setHorizontalHeaderLabels(['Upload Date', 'Part Number', 'Serial Number', 'KPC Number', 'Measurement'])
        self.treeView.setModel(self.proxyModel)
        self.treeView.setSortingEnabled(True)
        
            
        layout.addWidget(self.treeView, 4, 0, 1, 6)
        
        
        self.setLayout(layout)
            
        self.close()
        
    def getTolerance(self, selectedPartData, kpcNum):
        for feature in selectedPartData['features']:
            if feature['kpcNum'] == kpcNum:
                return feature['tol']
        return "N/A"
    
    def deleteMeasurement(self):
        index = self.treeView.currentIndex()
        if not index.isValid():
            QMessageBox.warning(self, "Error", "No measurement selected.")
            return
        
        sourceIndex = self.proxyModel.mapToSource(index)
        uploadDate = self.model.itemFromIndex(sourceIndex.siblingAtColumn(0)).text()
        stripped_date = datetime.strptime(uploadDate, '%m/%d/%Y')
        formatted_date = datetime.strftime(stripped_date, '%m/%d/%y')
        serialNumber = self.model.itemFromIndex(sourceIndex.siblingAtColumn(1)).text()
        
        if uploadDate and serialNumber:
            partNumber = self.partNumber.text()
            result = database.delete_measurement_by_id(self, partNumber, serialNumber, formatted_date)
            if result > 0:
                self.currentRow = sourceIndex.row()
                self.refreshTreeView()
        else: 
            QMessageBox.warning(self, "Error", "Could not retrieve data.")
    
    def loadPartData(self, selectedPartData, selectedPartUploadData):
        self.partNumber.setText(selectedPartData['partNumber'])
        self.revLetter.setText(selectedPartData['rev'])
        
        self.model.clear()
        self.model.setHorizontalHeaderLabels(['Upload Date','Serial Number', 'KPC Number', 'Blueprint Requirement', 'Measurement'])
        self.treeView.sortByColumn(0, Qt.AscendingOrder)
        
        for uploadData in selectedPartUploadData:
            uploadDate = uploadData['uploadDate']
            formatted_date = format_date(uploadDate)
            serialNumber = uploadData['serialNumber']
            
            parentRow = [
                QStandardItem(formatted_date),
                QStandardItem(serialNumber),
                QStandardItem(""),
                QStandardItem("")
            ]
            self.model.appendRow(parentRow)
            
            if 'measurements' in uploadData:
            
                for measurement in uploadData['measurements']:
                    kpcNum = measurement['kpcNum']
                    meas = measurement['measurement']
                    tolerance = self.getTolerance(selectedPartData, kpcNum)
                    childRow = [
                        QStandardItem(""),
                        QStandardItem(""),
                        QStandardItem(kpcNum),
                        QStandardItem(tolerance),
                        QStandardItem(meas)
                    ]
                    parentRow[0].appendRow(childRow)
        self.model.layoutChanged.emit()
                    
    def refreshTreeView(self):
        part_id = self.partId
        selectedPartData = database.get_part_by_id(part_id)
        selectedPartUploadData = database.get_measurements_by_id(part_id)
        currentSortColumn = self.treeView.header().sortIndicatorSection()
        currentSortOrder = self.treeView.header().sortIndicatorOrder()
        
        self.loadPartData(selectedPartData, selectedPartUploadData)
        
        if hasattr(self, 'currentRow'):
            rowCount = self.model.rowCount()
            if rowCount > 0:
                newRowIndex = min(self.currentRow, rowCount - 1)
                newIndex = self.model.index(newRowIndex, 0)
                mappedIndex = self.proxyModel.mapFromSource(newIndex)
                self.treeView.setCurrentIndex(mappedIndex)
                print(mappedIndex.row(), mappedIndex.column())
                self.treeView.scrollTo(mappedIndex, 5)
        
        self.treeView.header().setSortIndicator(currentSortColumn, currentSortOrder)
        self.treeView.sortByColumn(currentSortColumn, currentSortOrder)
        
    def closeWindow(self):
        self.close()
        
class DateSortProxyModel(QSortFilterProxyModel):
    def __init__(self, dateColumnIndex, *args, **kwargs):
        super().__init__( *args, **kwargs)
        self.dateColumnIndex = dateColumnIndex
    def lessThan(self, left, right):
        if left.column() == self.dateColumnIndex and right.column() == self.dateColumnIndex:
            leftData = self.sourceModel().data(left)
            rightData = self.sourceModel().data(right)
        
            leftDate = QDate.fromString(leftData, 'MM/dd/yyyy')
            rightDate = QDate.fromString(rightData, 'MM/dd/yyyy')
        
            
            return leftDate < rightDate
        else:
            return super(DateSortProxyModel, self).lessThan(left, right)
        
def renderDashboard():
    global d
    d = dashboard()
    d.show()
    window.close()
    
def login(email, pwd):
        renderDashboard()
    # auth = pwd.encode()
    # password = hashlib.md5(auth).hexdigest()
    # storedData = cursor.execute("SELECT * FROM user WHERE email = (?) AND password = (?)", [email, password])
    
    # loginQueue = []
    
    # for row in storedData:
    #     for x in row:
    #         loginQueue.append(x)
            
    # if (email and password) in loginQueue:
    #     renderDashboard()
    # else:
    #     loginFailed()
        
def register(email, pwd, confPwd):
    if confPwd != pwd:
        passMismatch()
    elif(re.fullmatch(emailRegex, email)) and(re.fullmatch(passwordRegex, pwd)):
        enc = confPwd.encode()
        password = hashlib.md5(enc).hexdigest()
        
        loginQueue = []
    
        if email in loginQueue:
            userExists()
        else:
            print('user has been registered')
        
    else:
        invalid()
        
def format_date(date_str):
    parts = date_str.split('/')
    if len(parts) == 3:
        month, day, year = parts
        if len(year) == 2:
            date_obj = datetime.strptime(date_str, '%m/%d/%y')
        elif len(year) == 4:
            date_obj = datetime.strptime(date_str, '%m/%d/%Y')
        else:
            raise ValueError('Invalid date format')
            
        return date_obj.strftime('%m/%d/%Y')
    else: 
        raise ValueError('Invalid date format')

def loginFailed():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('Login Failed')
    dlg.exec()
    
def userExists():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('User Already Exists')
    dlg.exec()
    
def invalid():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('Invalid Email or Password')
    dlg.exec()
    
def passMismatch():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('Passwords do not match')
    dlg.exec()
    
def parse_tolerance(tolerance):
    range_pattern = re.compile(r'(?<!\S)(\d*\.\d+)\s*-\s*(\d*\.\d+)(?!\S)')
    specific_tolerance_pattern = re.compile(r'([A-Za-z ]+)\s+(\d*\.\d+)')
    
    range_match = range_pattern.search(tolerance)
    
    if range_match:
        max_val, min_val = map(float, range_match.groups())
        return ( max_val, min_val)
    
    specific_tolerance_match = specific_tolerance_pattern.search(tolerance)
    if specific_tolerance_match:
        tolerance_type, value = specific_tolerance_match.groups()
        return (0, float(value))
    
    return None, None

def calculate_cpk(data, usl=None, lsl=None, target=None):
    if not data or len(data) < 2 or np.isnan(data).any() or np.isinf(data).any():
        return None
    sigma = np.std(data, ddof=1)
    mean = np.mean(data)
    
    if sigma == 0:
        return None
    
    cpk = None
    if usl is None and lsl is not None:
        cpk = (mean - lsl) / (3 * sigma)
        return cpk
    elif usl is not None and lsl is not None:
        cpk_upper = (usl - mean) / (3 * sigma)
        cpk_lower = (mean - lsl) / (3 * sigma)
        cpk = min(cpk_upper, cpk_lower)
    elif usl is not None:
        cpk = (usl - mean) / (3 * sigma)
    elif lsl is not None:
        cpk = (mean - lsl) / (3 * sigma)
    
    if target is not None:
        if usl is not None:
            cpk = (usl - target) / (3 * sigma)
        elif lsl is not None:
            cpk = (target - lsl) / (3 * sigma)
    return cpk
    
app = QApplication(sys.argv)
window = loginWindow()
window.show()
app.exec_()