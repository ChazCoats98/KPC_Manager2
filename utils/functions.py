import re
import os
import tempfile
import shutil
import hashlib
import win32com.client
import time
from pyqtspinner import WaitingSpinner
from datetime import (
    datetime, 
    timedelta, 
    date
    )
import numpy as np
from scipy import interpolate
from scipy.stats import anderson, shapiro, boxcox, norm, lognorm, expon, weibull_min, genextreme, gamma, logistic,fisk, kstest, weibull_max, gumbel_l, gumbel_r
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMessageBox
from utils import database
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

emailRegex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
passwordRegex = r'^(?=.*?[A-Z])(?=.*?[a-z])(?=.*?[0-9])(?=.*?[#?!@$%^&*-]).{8,}$'

def login(email, pwd):
        renderDashboard()
    # auth = pwd.encode()
    # password = hashlib.md5(auth).hexdigest()
    # storedData = cursor.execute("SELECT * FROM user WHERE email = (?) AND password = (?)", [email, password])
    
    # loginQueue = []
    
    # for row in storedData:
    #     for x in row:
    #         loginQueue.append(x)
            
    # if (email and password) in loginQueue:
    #     renderDashboard()
    # else:
    #     loginFailed()
        
def register(email, pwd, confPwd):
    if confPwd != pwd:
        passMismatch()
    elif(re.fullmatch(emailRegex, email)) and(re.fullmatch(passwordRegex, pwd)):
        enc = confPwd.encode()
        password = hashlib.md5(enc).hexdigest()
        
        loginQueue = []
    
        if email in loginQueue:
            userExists()
        else:
            print('user has been registered')
        
    else:
        invalid()
        
def format_date(date_str):
    parts = date_str.split('/')
    if len(parts) == 3:
        month, day, year = parts
        if len(year) == 2:
            date_obj = datetime.strptime(date_str, '%m/%d/%y')
        elif len(year) == 4:
            date_obj = datetime.strptime(date_str, '%m/%d/%Y')
        else:
            raise ValueError('Invalid date format')
            
        return date_obj.strftime('%m/%d/%Y')
    else: 
        raise ValueError('Invalid date format')
    
    
#_____________________________#
##Ppap Form View Functions##
#_____________________________#
#_____________________________#
##Part Form View Functions##
#_____________________________#

#Adds new part or updates current selected part if in edit mode
def submitPart(self):
        upload_date_str = self.udInput.text().strip()
        if not upload_date_str:
            QMessageBox.warning(self, "Error", "Upload date cannot be empty.")
            return
        try: 
            upload_date = datetime.strptime(upload_date_str, '%m/%d/%Y')
        except ValueError:
            QMessageBox.warning(self, "Error", "Upload date must be in MM/DD/YYYY Format.")
            return
            
        due_date = upload_date + timedelta(days=90)
        due_date_str = due_date.strftime('%m/%d/%Y')
        new_part_data = {
            "partNumber": self.partInput.text(),
            "rev": self.revInput.text(),
            "uploadDate": self.udInput.text(),
            "dueDate": due_date_str,
            "notes": self.notesInput.text(),
            "currentManufacturing": self.manufacturingCheck.isChecked(),
            "features": []
        }
        for row in range(self.featureTable.rowCount()):
            feature = {
                "feature": self.featureTable.item(row, 0).text(),
                "designation": self.featureTable.item(row, 1).text(),
                "kpcNum": self.featureTable.item(row, 2).text(),
                "opNum": self.featureTable.item(row, 3).text(),
                "tol": self.featureTable.item(row, 4).text(),
                "engine": self.featureTable.item(row, 5).text(),
            }
            new_part_data["features"].append(feature)
        def on_submit_success(is_success):
            if is_success:
                self.partSubmitted.emit()
        if self.mode == "add":
            if database.check_for_part(self.partInput.text()):
                QMessageBox.warning(self, "Error", "Part already exists in database.")
                return
            else: 
                database.submit_new_part(new_part_data, callback=on_submit_success)
        elif self.mode == "edit":
            database.update_part_by_id(self.partId, new_part_data, callback=on_submit_success)
            
        self.close()
        
#_____________________________#
##Ppap Form View Functions##
#_____________________________#
#Adds new ppap part or updates current selected part if in edit mode
def submitPPAPPart(self):
        partNum = self.partInput.text().strip()
        revLetter = self.revInput.text().strip()
        ppapNum = self.ppapInput.text().strip()
        ppapPhase = self.phaseInput.text().strip()
        if self.intBCheck.isChecked():
            intBDate = None
        else:
            intBDate = self.intBBox.text()
            
        if self.intACheck.isChecked():
            intADate = None
        else:
            intADate = self.intABox.text()
        if self.fullCheck.isChecked():
            fullDate = None
        else:
            fullDate = self.fullBox.text()
        
        
        if not partNum:
            QMessageBox.warning(self, "Error", "Please enter a valid part number")
            return
        elif not revLetter:
            QMessageBox.warning(self, "Error", "Please enter a valid revision letter")
            return
        elif not ppapNum:
            QMessageBox.warning(self, "Error", "Please enter a valid PPAP package number")
            return
        elif not ppapPhase:
            QMessageBox.warning(self, "Error", "Please enter the current PPAP phase")
            return
            
        ppap_part_data = {
            "partNumber": partNum,
            "rev": revLetter,
            "ppapNumber": ppapNum,
            "ppapPhase": ppapPhase,
            "intBDate": intBDate,
            "intADate": intADate,
            "fullDate": fullDate,
            "elements": []
        }
        for row in range(self.elementsTable.rowCount()):
            element = {
                "element": self.elementsTable.item(row, 0).text(),
                "document": self.elementsTable.item(row, 1).text(),
                "submitted": self.elementsTable.cellWidget(row, 2).isChecked(),
                "submitDate": self.elementsTable.item(row, 3).text(),
                "status": self.elementsTable.item(row, 4).text() if self.elementsTable.item(row, 4) else 'Initial',
                "interimB": self.elementsTable.item(row, 5).text() if self.elementsTable.item(row, 5) else 'In Process',
                "interimA": self.elementsTable.item(row, 6).text() if self.elementsTable.item(row, 6) else 'In Process',
                "full": self.elementsTable.item(row, 7).text() if self.elementsTable.item(row, 7) else 'In Process',
                "notes": self.elementsTable.item(row, 8).text() if self.elementsTable.item(row, 8) else None
            }
            ppap_part_data["elements"].append(element)
        print(ppap_part_data)
        def on_submit_success(is_success):
            if is_success:
                self.ppapSubmitted.emit()
        if self.mode == "add":
            if database.check_for_ppap(self.partInput.text()):
                QMessageBox.warning(self, "Error", "Part already exists in database.")
                return
            else: 
                database.submit_new_ppap_part(ppap_part_data, callback=on_submit_success)
        elif self.mode == "edit":
            database.update_ppap_by_id(self.partId, ppap_part_data, callback=on_submit_success)
            
        self.close()
        
#Loads part features to table for editing
def addFeatureToTable(self, feature_data):
        row_position = self.featureTable.rowCount()
        self.featureTable.insertRow(row_position)
        
        keys = ['feature', 'designation', 'kpcNum', 'opNum', 'tol', 'engine']
        for i, key in enumerate(keys):
            value = feature_data.get(key, '')
            self.featureTable.setItem(row_position, i, QTableWidgetItem(value))
            
def addFeatureToFormTable(self, feature_data):
        row_position = self.featureTable.rowCount()
        self.featureTable.insertRow(row_position)
        
        keys = ['feature', 'designation', 'kpcNum', 'opNum', 'tol', 'cpk', 'formNumber', 'uploadDate', 'ms2Date', 'ms3Date', 'ms4Date']
        for i, key in enumerate(keys):
            value = feature_data.get(key, '')
            checkbox = QCheckBox()
            checkbox.setCheckState(False)
            self.featureTable.setCellWidget(row_position, 0, checkbox)
            self.featureTable.setItem(row_position, i + 1, QTableWidgetItem(str(value)))
            
#_____________________________#
##KPC Snapshot View Functions##
#_____________________________#
            
def addKPCToTable(self):
        self.kpcTable.setRowCount(0)
        data = database.get_all_data()
        
        features = []
        
        for part in data:
            for feature in part['features']:
                if 'cpk' in feature:
                    if feature['cpk'] < 1.33:
                        features.append((
                            part['partNumber'], 
                            feature['kpcNum'], 
                            feature['tol'], 
                            part['uploadDate'], 
                            feature['cpk']
                        ))
                
        features.sort(key=lambda x: x[4])
        
        for feature in features:
            row_position = self.kpcTable.rowCount()
            self.kpcTable.insertRow(row_position)
            
            for col, data in enumerate(feature):
                item = QTableWidgetItem(str(data))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable | ~Qt.ItemIsSelectable)
                self.kpcTable.setItem(row_position, col, item)
                
def setMinHeaderWidth(self):
    total_width = self.kpcTable.width()
    
    for column in range(self.kpcTable.columnCount()):
        self.kpcTable.horizontalHeader().setSectionResizeMode(column, QHeaderView.ResizeToContents)
    
    content_width = sum(self.kpcTable.columnWidth(column) for column in range(self.kpcTable.columnCount()))
    extra_width = total_width - content_width
    
    if extra_width > 0:
        for column in range(self.kpcTable.columnCount()):
            current_width = self.kpcTable.columnWidth(column)
            
            proportion = current_width / content_width
            new_width = current_width + int(proportion * extra_width)
            self.kpcTable.setColumnWidth(column, new_width)
    
            

#_____________________________#
##Upload Data View Functions##
#_____________________________#

def create_spinner(self):
    parent = self
    spinner = WaitingSpinner(
    parent,
    roundness=100.0,
    opacity=10.32,
    fade=53.87,
    radius=15,
    lines=107,
    line_length=20,
    line_width=6,
    speed=0.75,
    color=(60, 145, 235)
)


#clears lot inputs from table 
def clearLotInputs(self):
        while self.scrollAreaWidgetLayout.count():
            item = self.scrollAreaWidgetLayout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.serialNumberInputs = []
        self.featureTables = []
        
def loadMeasurementData(self, measurement_data):
    createLotInputs(self, len(measurement_data))
    
    for i, measurements in enumerate(measurement_data):
        if i >= len(self.featureTables):
            break
        
        feature_table = self.featureTables[i]
        measurement_map = {m['kpcNum']: m['measurement'] for m in measurements['measurements']}
        self.serialNumberInputs[i].setText(measurements["serialNumber"])
        
        for row in range(feature_table.rowCount()):
            kpc_num = feature_table.item(row, 1).text()
            if kpc_num in measurement_map:
                measurement_value = measurement_map[kpc_num]
                feature_table.setItem(row, 4, QTableWidgetItem(measurement_value))
        
        
def filterData(self, part_number):
    measurement_data = database.get_measurements_by_id(part_number)
    measurement_data.sort(key= lambda doc: datetime.strptime(doc["uploadDate"], "%m/%d/%Y"))

#Create tables for data upload    
def createLotInputs(self, lot_size):
        currentInputs = {
            'serialNumbers': [input.text() for input in self.serialNumberInputs],
            'measurements': []
        }
        
        for table in self.featureTables:
            currentTableData = []
            for row in range(table.rowCount()):
                rowData = [table.item(row, col).text() if table.item(row, col) else '' for col in range(table.columnCount())]
                currentTableData.append(rowData)
            currentInputs['measurements'].append(currentTableData)
        
        
        clearLotInputs(self)
        
        partData = database.get_part_by_id(self.partId)
        if not partData:
            QMessageBox.warning(self, "Error", "Part data not found in database.")
            return
        part_features = partData['features']
        
        for lot_index in range(lot_size):
            serialNumberInput = QLineEdit()
            serialNumberInput.setPlaceholderText(f'Enter Serial Number {lot_index+1}')
            serialNumberInput.textChanged.connect(lambda text, sn_input=serialNumberInput: checkSerialNumber(self, text, sn_input))
            self.serialNumbersLayout.addWidget(serialNumberInput)
            self.serialNumberInputs.append(serialNumberInput)
            
            featureTable = QTableWidget()
            featureTable.setColumnCount(5)
            featureTable.setRowCount(len(part_features))
            featureTable.setHorizontalHeaderLabels(['Feature Number', 'KPC Number', 'Blueprint Dimension', 'Op Number', 'Measurement'])
            featureTable.horizontalHeader().setStretchLastSection(False)
            for column in range(featureTable.columnCount()):
                featureTable.horizontalHeader().setSectionResizeMode(column, QHeaderView.Stretch)
                
            featureTable.setFixedHeight(250)
                
            for row, feature in enumerate(part_features):
                featureTable.setItem(row, 0, QTableWidgetItem(feature['feature']))
                featureTable.setItem(row, 1, QTableWidgetItem(feature['kpcNum']))
                featureTable.setItem(row, 2, QTableWidgetItem(feature['tol']))
                featureTable.setItem(row, 3, QTableWidgetItem(feature.get('opNum', '')))
                featureTable.setItem(row, 4, QTableWidgetItem(''))
                
            
            adjustTableHeight(featureTable)
            self.featureTables.append(featureTable)
            
            self.scrollAreaWidgetLayout.addWidget(serialNumberInput)
            self.scrollAreaWidgetLayout.addWidget(featureTable)
            
        if len(currentInputs['serialNumbers']) <= lot_size:
            for i, text in enumerate(currentInputs['serialNumbers']):
                self.serialNumberInputs[i].setText(text)
        else:    
            for i, text in enumerate(currentInputs['serialNumbers'][:lot_size]):
                self.serialNumberInputs[i].setText(text)
                
        if len(currentInputs['measurements']) <= lot_size:
            for i, table_data in enumerate(currentInputs['measurements']):
                for row, row_data in enumerate(table_data):
                    for col, value in enumerate(row_data):
                        self.featureTables[i].setItem(row, col, QTableWidgetItem(value))
        else:
            for i, table_data in enumerate(currentInputs['measurements'][:lot_size]):
                for row, row_data in enumerate(table_data):
                    for col, value in enumerate(row_data):
                        self.featureTables[i].setItem(row, col, QTableWidgetItem(value))
                        
#Adjusts table height for lot creation function
def adjustTableHeight(table):
        total_height = table.horizontalHeader().height()
        for i in range(table.rowCount()):
            total_height += table.rowHeight(i)
        
        margin = 4
        total_height += margin
        
        table.setFixedHeight(total_height)

#Checks user input serial number against database to prevent duplicates
def checkSerialNumber(self, text, sender):
        exists = database.check_serial_number(text)
        if exists:
            sender.setStyleSheet("background-color: red")
            QMessageBox.warning(self, "Serial Number Invalid", "Data for this serial number has already been uploaded.")
        else: 
            sender.setStyleSheet("background-color: white")
            
#opens file dialog for selecting pdf to extract data from
def openPdfFileDialog(self):
        partNumber = self.partNumber.text()
        initialDir = f'//Server/d/Inspection/CMM Files/Printouts/{partNumber}'
            
        filePath, _ = QFileDialog.getOpenFileName(
            self, 
            "Open PDF File", 
            initialDir, 
            "PDF Files (*.pdf)"
        )
        if filePath:
            extractDataFromPdf(self, filePath)
#Lets user select a PDF to pull measurement data from. Not functional yet.
#issues with PDF structure and correctly identifying actual measurements.
#Plan to fix with ML model
def extractDataFromPdf(self, filePath):
        for page_layout in extract_pages(filePath):
            for element in page_layout:
                if isinstance(element, LTTextContainer):
                    for text_line in element:
                        text = text_line.get_text().strip()
                        print(text)
        #part_data = database.get_part_by_id(self.partNumber.text())
        #tolerances = {feature['kpcNum']: parse_tolerance(feature['tol']) for feature in part_data['features']}
        #print(tolerances)
        #try:
            #reader = PdfReader(filePath)
            #text = ''
            #for page in reader.pages:
                #text += page.extract_text() + '\n'
            
            #pattern = re.compile(r'\b\d*\.?\d+\b')
            #measurements = [float(m) for m in pattern.findall(text)]
            
            #matched_measurements = {kpc: [] for kpc in tolerances.keys()}
            #for measurement in measurements:
                #for kpc, tolerance in tolerances.items():
                    #lower, upper = tolerance
                    #if lower is not None and upper is not None and lower < measurement < upper:
                        # matched_measurements[kpc].append(measurement)
                    
            #print(matched_measurements)
            #for row in range(self.dataTable.rowCount()):
                #kpcNum = self.dataTable.item(row, 1).text()
                #if kpcNum in matched_measurements and matched_measurements[kpcNum]:
                    #self.dataTable.setItem(row, 4, QTableWidgetItem(str(matched_measurements[kpcNum][0])))
        #except Exception as e:
            #print(str(e))
            #QMessageBox.critical(self, "Error", f"Failed to read PDF: {str(e)}")

#Submits data to database and writes data to Net-Inspect template
def submitData(self):
        part_number = self.partNumber.text()
        existing_part_data = database.get_part_by_id(part_number)
        measurement_data = database.get_measurements_by_id(part_number)
        machine = self.machineComboBox.currentText()
        run_number = self.runNumberInput.text()
        lot_size = self.lotSizeComboBox.currentText()
        template_path = './utils/Templates/Measurement_Import_template.xlsx'
        today = datetime.today()
        upload_date_str = today.strftime('%m/%d/%Y')
        upload_date_file_path = today.strftime('%m-%d-%Y')
        new_file_path = f'./Results/{part_number}_data_upload_{upload_date_file_path}.xlsx'
        due_date = today + timedelta(days=90)
        due_date_str = due_date.strftime('%m/%d/%Y')
        today = datetime.today()
        upload_date_str = today.strftime('%m/%d/%Y')
        part_number = self.partNumber.text()
        machine = self.machineComboBox.currentText()
        run_number = self.runNumberInput.text()
        lot_size = self.lotSizeComboBox.currentText()
        upload_date_file_path = today.strftime('%m-%d-%Y')
        template_path = './utils/Templates/Measurement_Import_template.xlsx'
        new_file_path = f'./Results/{part_number}_data_upload_{upload_date_file_path}.xlsx'
        due_date = today + timedelta(days=90)
        due_date_str = due_date.strftime('%m/%d/%Y')

        if not existing_part_data:
            QMessageBox.warning(self, "Error", "Part data not found in database.")
            return
        
        
        if len(run_number) == 0:
            QMessageBox.warning(self, "Error", "Please enter a valid run number")
            return
        elif len(machine) == 0:
            QMessageBox.warning(self, "Error", "Please select a valid machine")
            return
        elif len(lot_size) == 0:
            QMessageBox.warning(self, "Error", "Please select a lot size")
            return
        
        shutil.copy(template_path,  new_file_path)
        
        upload_data, updated_part_data = generateExcel(
            self,
            part_number=part_number,
            serial_number_inputs=self.serialNumberInputs,
            feature_tables=self.featureTables,
            machine=machine,
            run_number=run_number,
            lot_size=lot_size,
            upload_date_file_path=upload_date_file_path
            )
        
        try: 
            database.add_measurement(upload_data)
            
            def on_submit_success(is_success):
                if is_success:
                    self.dataSubmitted.emit()
                    QMessageBox.information(self, "Success", "Data uploaded successfully.")
                    clearLotInputs(self)
            for part in updated_part_data:
                database.update_part_by_id(self.partId, part, callback=on_submit_success)
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred while saving the file: {e}")
        
        cpk_values = calculateCpk(self, measurement_data)
        database.save_cpk_values(part_number, cpk_values)
            
            
def generateExcel(self, part_number, serial_number_inputs, feature_tables, machine, run_number, lot_size, upload_date_file_path):
    target_row = 2
    template_path = './utils/Templates/Measurement_Import_template.xlsx'
    new_file_path = f'./Results/{part_number}_data_upload_{upload_date_file_path}.xlsx'
    today = datetime.today()
    upload_date_str = today.strftime('%m/%d/%Y')
    due_date = today + timedelta(days=90)
    due_date_str = due_date.strftime('%m/%d/%Y')
    upload_data_list = []
    updated_part_data_list = []
    
    shutil.copy(template_path,  new_file_path)
    workbook = load_workbook(filename=new_file_path)
    sheet = workbook.active
        
    try:
        for serial_input, feature_table in zip(serial_number_inputs, feature_tables):
            serial_number = serial_input.text()
            if database.check_serial_number(serial_number):
                duplicate_warning = QMessageBox(self)
                duplicate_warning.setWindowTitle("Duplicate Serial Number")
                duplicate_warning.setText(f"Serial Number {serial_number} already in database. Would you like to save the data anyways?")
                duplicate_warning.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                duplicate_warning.setIcon(QMessageBox.Warning)
                    
                response = duplicate_warning.exec_()
                if response == QMessageBox.No:
                    break
                else:
                    continue
                
            upload_data = {
                "partNumber": part_number,
                "serialNumber": serial_number,
                "uploadDate": upload_date_str,
                "measurements": []
            }
                
            updated_part_data = {
                    "uploadDate": upload_date_str,
                    "dueDate": due_date_str,
                    "features" : []
            }
            
            for row in range(feature_table.rowCount()):
                feature_number = feature_table.item(row, 0).text()
                kpc_number= feature_table.item(row, 1).text()
                op_number = feature_table.item(row, 3).text()
                measurement = feature_table.item(row, 4).text()
            
                if part_number:
                    sheet.cell(row=target_row, column=1).value = part_number
                if feature_number:
                    sheet.cell(row=target_row, column=2).value = feature_number
                if machine:
                    sheet.cell(row=target_row, column=3).value = machine
                if run_number:
                    sheet.cell(row=target_row, column=4).value = run_number
                if lot_size:
                    sheet.cell(row=target_row, column=5).value = lot_size
                if measurement:
                    sheet.cell(row=target_row, column=6).value = measurement
                if serial_number:
                    sheet.cell(row=target_row, column=8).value = serial_number
                target_row += 1
                
                upload_data["measurements"].append({
                    "kpcNum": kpc_number,
                    "measurement": measurement
                })
                
                updated_part_data["features"].append({
                    "opNum": op_number if op_number else "",
                })
                
            upload_data_list.append(upload_data)
            updated_part_data_list.append(updated_part_data)
            
        workbook.save(filename=new_file_path)
    except Exception as e:
        print(e)
        QMessageBox.critical(self, "Error", f"An unexpected error occurred: {str(e)}")
        
    try: 
        basePath = f'//server/D/Quality Control/UPPAP Records/Process Cert + Data Collection/Data points/{part_number}/762396_reupload.xlsx'
        fileName, ok = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            basePath,
            "Excel files (*.xlsx)"
        )
            
    except Exception as e:
        print(e)
        QMessageBox.critical(self, "Error", f"An unexpected error occurred: {e}")
        
    if not fileName:
        QMessageBox.information(self, "Save Canceled", "The save operation was cancelled")
        return
        
    if fileName:
        if not fileName.endswith('.xlsx'):
            fileName += '.xlsx'
        try: 
            shutil.move(new_file_path, fileName)
            return upload_data_list, updated_part_data_list
        
        except Exception as e:
            print(e)
            QMessageBox.critical(self, "Error", f"An error occurred while saving the file: {e}")
            return None, None
            
def loadExcel(self, part_number):
    today = datetime.today()
    upload_date_file_path = today.strftime('%m-%d-%Y')
    upload_date_str = today.strftime('%m/%d/%Y')
    part_number = self.partNumber.text()
    machine = self.machineComboBox.currentText()
    run_number = self.runNumberInput.text()
    lot_size = len(self.serialNumberInputs)
    print(lot_size)
    serial_number_inputs = self.serialNumberInputs
    
    generateExcel(self, part_number=part_number,serial_number_inputs=serial_number_inputs,feature_tables=self.featureTables, machine=machine, run_number=run_number, lot_size=lot_size, upload_date_file_path=upload_date_file_path )


def loginFailed():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('Login Failed')
    dlg.exec()
    
def userExists():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('User Already Exists')
    dlg.exec()
    
def invalid():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('Invalid Email or Password')
    dlg.exec()
    
def passMismatch():
    dlg = QMessageBox()
    dlg.setWindowTitle('ERROR')
    dlg.setText('Passwords do not match')
    dlg.exec()
    
#calculate or recalculate CPK on data upload 
def calculateCpk(self, measurement_data):
        self.spinner.start()
        QCoreApplication.processEvents()
        
        self.thread = QThread()
        self.worker = Worker(self.partId, self, measurement_data)
        self.worker.moveToThread(self.thread)
        
        cpk_values = []
        
        def store_result(result):
            nonlocal cpk_values
            cpk_values = result
            
            
        self.thread.started.connect(self.worker.run)
        self.worker.init_dialog.connect(info_dialog_init)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker.finished.connect(self.spinner.stop)
        self.worker.result_ready.connect(store_result)
        self.thread.finished.connect(self.thread.deleteLater)   
        event_loop = QEventLoop()
        self.worker.result_ready.connect(event_loop.quit)
        self.thread.start()
        
        event_loop.exec_()
        
        return cpk_values
    


#Worker to task out long-running cpk calculation to prevent application freeze
class Worker(QObject):
    finished = pyqtSignal()
    init_dialog = pyqtSignal(QWidget, str, str)
    result_ready = pyqtSignal(list)
    
    def __init__(self, partId, parent = None, measurement_data = None):
        super().__init__()
        self.partId = partId
        self.parent = parent
        self.measurement_data = measurement_data
        
    def run(self): 
        formatted_cpk_values = self.calculateAndUpdateCpk()
        if formatted_cpk_values:
            self.result_ready.emit(formatted_cpk_values)
            self.finished.emit()
        
    def calculateAndUpdateCpk(self):
        print("Worker: Starting calculation")
        partId = self.partId
        part_data = database.get_part_by_id(partId)
        print(part_data)
        if part_data:
            tolerances = {feature['kpcNum']: self.parse_tolerance(feature.get('tol', '0-0')) for feature in part_data.get('features', []) if feature['kpcNum']}
            
    
        
        measurements_by_kpc = {kpc: [] for kpc in tolerances.keys()}
        
        if self.measurement_data:
            for entry in self.measurement_data:
                for measurement in entry.get('measurements', []):
                    kpcNum = measurement.get('kpcNum')
                    if kpcNum and kpcNum in measurements_by_kpc:
                        if measurement['measurement']:
                            measurements_by_kpc[kpcNum].append(float(measurement['measurement']))

        for kpc, measurements in measurements_by_kpc.items():
            if len(measurements) <= 3:
                self.init_dialog.emit(self.parent, 'Too Few Data Points', 'Not enough data points. Skipping CPK calculation')
                return
        
        
        dist_data, percentiles = self.calculate_dist(measurements_by_kpc, tolerances)
        if dist_data:
            cpk_values = {}
            print('calculating cpk')
            for kpc, data in dist_data.items():
                if data['dist'] == 'Normal':
                    lsl, usl = tolerances[kpc]
                    measurement = measurements_by_kpc[kpc]
                    sigma = np.std(measurement, ddof=1)
                    mean = np.mean(measurement)
                
                    if usl is None and lsl is not None:  
                        cpl = (mean - lsl) / (3 * sigma)
                        cpk = cpl  
                    elif lsl is None and usl is not None: 
                        cpu = (usl - mean) / (3 * sigma)
                        cpk = cpu  
                    elif usl is not None and lsl is not None:  
                        if usl < lsl:
                            usl, lsl = lsl, usl
                        cpu = (usl - mean) / (3 * sigma)
                        cpl = (mean - lsl) / (3 * sigma)
                        cpk = min(cpl, cpu)
                
                    cpk = min(cpl, cpu)
                    print(f'cpk: {cpk}')
                    cpk_values[kpc] = cpk
                else: 
                    calc_type, ppk = self.calculate_ppk(percentiles[kpc], tolerances[kpc])
                    print(f'ppk: {ppk}')
                    cpk_values[kpc] =  ppk
        
                print(cpk_values)
            print("Worker: Calculation finished")
            if cpk_values:
                tol_values = {feature['kpcNum']: feature.get('tol', 'N/A') for feature in part_data.get('features', []) if 'kpcNum' in feature}
                formatted_cpk_values = []
                for kpc, cpk in cpk_values.items():
                    formatted_cpk_values.append({
                        "kpcNum":kpc,
                        "cpk": round(abs(cpk if cpk is not None else 0), 3),
                        "tol": tol_values.get(kpc, 'N/A') 
                })
                return formatted_cpk_values
        else:
            return {}
    
    def test_normalRJ(self, measurements, tolerances):
        normality_results = {}
        for kpc, measurements in measurements.items():
            
            stat, p_value = shapiro(measurements)
            
            normality_results[kpc] = {
                'Shapiro Stat': stat,
                'shapiro P-value': p_value
            }
        return normality_results

    def calculate_dist(self, measurements, tolerances):
        print("measurements: ")
        print(measurements)
        print("Tolerances: ")
        print(tolerances)
        try:
            mtb = win32com.client.Dispatch("Mtb.Application.1")
            mtb.UserInterface.Visible = False
            project = mtb.ActiveProject
            worksheet = project.ActiveWorksheet
            columns = worksheet.Columns
    
            dist_data = {}
            p_res = {}
            d = 1
            for i, (kpc, data) in enumerate(measurements.items()):
                usl, lsl = tolerances[kpc]
                if usl is not None and lsl is not None:
                    if usl < lsl:
                        usl, lsl = lsl, usl
                elif usl is None and lsl is not None:
                    usl = max(data) 
                elif lsl is None and usl is not None:
                    lsl = min(data) 
                else:
                    continue
                
                column = columns.Add(None, None, 1)
                column.SetData(data)
            
                command = f"DCapa C{d} 1; All; BoxCox; Johnson 0.10; RDescriptive; RFitTests; REstimate."
                project.ExecuteCommand(command)
    
                time.sleep(2)

                commands = project.Commands
                lastCommand = commands.Item(commands.Count)
                outputs = lastCommand.Outputs

                results = []

                for i in range(1, outputs.Count +1):
                    output = outputs.Item(i)
                    results.append(output.Text)
    
                formattedResults = "\n".join(results)
            
                parsed_gof = self.parse_goodness_of_fit(formattedResults)
                parsed_params = self.parse_distribution_params(formattedResults)
            
                best_fit = self.determine_best_fit(parsed_gof)
                best_fit_params = parsed_params[best_fit]
            
                percentiles = [.00135, .5, .99865]
                perc_dict = {}
                for percentile in percentiles:
                    p_column = columns.Add(None, d, 1)
                    p_column.SetData(percentile)
                    percent = self.get_perc_format(d, best_fit, best_fit_params)
                    project.ExecuteCommand(percent)
                    o_column = columns.Item(d+2).GetData(1)
                    perc_dict[f'{(percentile * 100):.3f}th Percentile'] = o_column
                    d = d + 2
                p_res[kpc] = perc_dict
                d = d + 1
    
                dist_data[kpc] = {
                    'dist': best_fit,
                    'params': best_fit_params
                }
            temp_file = os.path.join(tempfile.gettempdir(), 'temp_project.mpjx')
            project.SaveAs(temp_file)

            mtb.Quit()
            return dist_data, p_res
        except Exception as e:
            self.init_dialog.emit(self.parent, 'Error', f'Error trying to connect to Minitab. {e}')
            return None, None

    def parse_goodness_of_fit(self, output):
        if output:
            gof_start_match = re.search(r'Distribution\s+AD\s+P(?:\s+LRT P)?', output)
            if not gof_start_match:
                print("Goodness of Fit section not found.")
                return {}
            gof_start = gof_start_match.end()
            
            gof_end_match = re.search(r"ML Estimates of Distribution Parameters", output)
            if not gof_end_match:
                print("ML Estimates section not found.")
                return {}
            
            gof_end = gof_end_match.start()
            gof_table = output[gof_start:gof_end].strip()
            parsed_data = {}
            
            for line in gof_table.strip().split("\n"):
                parts = re.split(r'\s{2,}', line)
                print(parts)
        
                distribution = parts[0].strip()
                AD, P, LRT_P = None, None, None
                
                if len(parts) == 2:  
                    AD, P = float(parts[1]), None
                elif len(parts) == 3:
                    AD, P = float(parts[1]), parts[2].strip()
                    P = float(P) if P.replace('.', '', 1).isdigit() else P
                elif len(parts) == 4:
                    AD, P, LRT_P = float(parts[1]), parts[2].strip(), parts[3].strip()
                    P = float(P) if P.replace('.', '', 1).isdigit() else P
                    LRT_P = float(LRT_P) if LRT_P.replace('.', '', 1).isdigit() else LRT_P
            
                if isinstance(P, str) and (P.startswith('>') or P.startswith('<')):
                    P = float(P[1:])
                if isinstance(LRT_P, str) and (LRT_P.startswith('>') or LRT_P.startswith('<')):
                    LRT_P = float(LRT_P[1:])
            
                parsed_data[distribution] = {
                    'AD': AD,
                    'P': P,
                    'LRT_P': LRT_P,
                }
        
                unwanted_dist = ['Box-Cox Transformation', 'Johnson Transformation', '3-Parameter Loglogistic', '3-Parameter Lognormal', 'Loglogistic', '3-Parameter Weibull', '3-Parameter Gamma']
                filtered_data = {dist: values for dist, values in parsed_data.items() if dist not in unwanted_dist}
        
            return filtered_data

    def parse_distribution_params(self, output):
        params_start = re.search(r"Distribution\s+Location\s+Shape\s+Scale\s+Threshold", output).end()
        params_table = output[params_start:].strip()
    
        parsed_data = []
        for line in params_table.strip().split('\n'):
            parts = re.split(r'\s{2,}', line)

            distribution = parts[0].strip().rstrip('*')
            location, shape, scale, threshold = None, None, None, None
        
            if distribution.startswith('Normal') or distribution.startswith('Box-Cox Transformation') or distribution.startswith('Lognormal') or distribution.startswith('Smallest Extreme Value') or distribution.startswith('Largest Extreme Value') or distribution.startswith('Logistic') or distribution.startswith('Loglogistic') or distribution.startswith('Johnson Transformation'):
                location, scale = float(parts[-2]), float(parts[-1])
            elif distribution.startswith('Exponential'):
                scale = float(parts[-1])
            elif distribution.startswith('2-Parameter Exponential'):
                scale, threshold = float(parts[-2]), float(parts[-1])
            elif distribution.startswith('Weibull'):
                shape, scale = float(parts[-2]), float(parts[-1])
            elif distribution.startswith('3-Parameter Weibull') or distribution.startswith('3-Parameter Gamma'):
                shape, scale, threshold = float(parts[-3]), float(parts[-2]), float(parts[-1])
            elif distribution.startswith('3-Parameter Loglogistic') or distribution.startswith('3-Parameter Lognormal'):
                location, scale, threshold = float(parts[-3]), float(parts[-2]), float(parts[-1])
            
            parsed_data.append((distribution, location, shape, scale, threshold))
    
    
        parsed_dict = {}
        for line in parsed_data:
            distribution = line[0].strip()
            location, shape, scale, threshold = line[1], line[2], line[3], line[4]
            parsed_dict[distribution] = {
                'location': location,
                'shape': shape,
                'scale': scale,
                'threshold': threshold
            }
        
        return parsed_dict

    def determine_best_fit(self, gof_results):
        best_fit = None
        best_ad = float('inf')
    
        for distribution, values in gof_results.items():
            ad_stat = values['AD']
            p_value = values['P']
            print(p_value)
        
            if ad_stat < best_ad:
                best_ad = ad_stat
                best_fit = distribution
            
            elif ad_stat == best_ad:
                if isinstance(p_value, float):
                    best_p_value = gof_results[best_fit]['P']
                    if isinstance(best_p_value, float) and p_value > best_p_value:
                        best_fit = distribution
        return best_fit

    #Simple math for PPK calculations
    def calculate_ppk(self, percentiles, tolerances):
        print(percentiles)
        usl, lsl = tolerances
        if usl is not None and lsl is not None:
            if usl < lsl:
                usl, lsl = lsl, usl
            
        low = percentiles['0.135th Percentile']
        med = percentiles['50.000th Percentile']
        high = percentiles['99.865th Percentile']
        print(low, med, high)
        if usl is None and lsl is not None:
            ppl = (med - lsl) / (med - low)
            ppu = None
        elif usl is not None and lsl is not None:
            ppu = (usl - med) / (high - med)
            ppl = (med - lsl) / (med - low)
        elif lsl is None and usl is not None:
            ppu = (usl - med) / (high - med)
            ppl = None
            
        if ppu is not None and ppl is not None:
            if ppu > ppl:
                calc_type = 'PPU'
                return calc_type, ppu
            else: 
                calc_type = 'PPL'
                return calc_type, ppl
        elif ppu is None:
            calc_type = 'PPL'
            return calc_type, ppl
        else:
            calc_type = 'PPU'
            return calc_type, ppu
        
    def parse_tolerance(self, tolerance):
        range_pattern = re.compile(r'(?<!\S)(\d*\.\d+)\s*-\s*(\d*\.\d+)(?!\S)')
        min_only_pattern = re.compile(r'(?<!\S)(\d*\.\d+)\s+(?:Min|Minimum)\s+\w+', re.IGNORECASE)
        specific_tolerance_pattern = re.compile(r'([A-Za-z ]+)\s+(\d*\.\d+)')
    
        range_match = range_pattern.search(tolerance)
    
        if range_match:
            max_val, min_val = map(float, range_match.groups())
            return ( max_val, min_val)
        
        min_only_match = min_only_pattern.search(tolerance)
        if min_only_match:
            min_val = float(min_only_match.group(1))
            return (min_val, None)
    
        specific_tolerance_match = specific_tolerance_pattern.search(tolerance)
        if specific_tolerance_match:
            tolerance_type, value = specific_tolerance_match.groups()
            return (0, float(value))
    
        return None, None
    
    def get_perc_format(self, d, best_fit, best_fit_params):
        print(f'Best Fit: {best_fit} \nbest fit params: {best_fit_params}')
        if best_fit == 'Largest Extreme Value':
            return f'InvCDF C{d + 1} C{d + 2}; Lextremevalue {best_fit_params['location']} {best_fit_params['scale']}.'
        elif best_fit == 'Smallest Extreme Value':
            return f'InvCDF C{d + 1} C{d + 2}; Sextremevalue {best_fit_params['location']} {best_fit_params['scale']}.'
        elif best_fit == 'Weibull':
            return f'InvCDF C{d + 1} C{d + 2}; {best_fit} {best_fit_params['shape']} {best_fit_params['scale']}.'
        elif best_fit == 'Exponential':
            return f'InvCDF C{d + 1} C{d + 2}; {best_fit} {best_fit_params['scale']} 0.'
        elif best_fit == '2-Parameter Exponential':
            return f'InvCDF C{d + 1} C{d + 2}; Exponential {best_fit_params['scale']} {best_fit_params['threshold']}.'
        else:
            return f'InvCDF C{d + 1} C{d + 2}; {best_fit} {best_fit_params['location']} {best_fit_params['scale']}.'
        
#Attempt at calculating the nth percentile using scipy. Not currently being used
#Using Minitab to calculate instead.
#Keeping incase i go back to this method
def calculate_percentile(dist_data):
    print(dist_data)
    results = {}
    for kpc, data in dist_data.items():
        distribution = data['dist']
        params = data['params']
        if distribution.startswith('Normal'):
            low = norm.ppf(.00135, loc=params['location'], scale=params['scale'])
            med = norm.ppf(.5, loc=params['location'], scale=params['scale'])
            high = norm.ppf(.99865, loc=params['location'], scale=params['scale'])
        elif distribution.startswith('Lognormal'):
            low = lognorm.ppf(.00135, params['scale'], loc=0, scale=params['location'])
            med = lognorm.ppf(.5, params['scale'], loc=0, scale=params['location'])
            high = lognorm.ppf(.99865, params['scale'], loc=0, scale=params['location'])
        elif distribution.startswith('3-Parameter Lognormal'):
            low = lognorm.ppf(.00135, s=1, loc=params['location'], scale=params['scale']) + params['location']
            med = lognorm.ppf(.5, s=1, loc=params['location'], scale=params['scale']) + params['location']
            high = lognorm.ppf(.99865, s=1, loc=params['location'], scale=params['scale']) + params['location']
        elif distribution.startswith('Exponential'):
            low = expon.ppf(.00135, loc=0, scale=params['scale'])
            med = expon.ppf(.00135, loc=0, scale=params['scale'])
            high = expon.ppf(.00135, loc=0, scale=params['scale'])
        elif distribution.startswith('2-Parameter Exponential'):
            low = expon.ppf(.00135, loc=params['threshold'], scale=params['scale'])
            med = expon.ppf(.5, loc=params['threshold'], scale=params['scale'])
            high = expon.ppf(.99865, loc=params['threshold'], scale=params['scale'])
        elif distribution.startswith('Weibull'):
            low = weibull_min.ppf(.00135, params['shape'], loc=0, scale=params['scale'])
            med = weibull_min.ppf(.5, params['shape'], loc=0, scale=params['scale'])
            high = weibull_min.ppf(.99865, params['shape'], loc=0, scale=params['scale'])
        elif distribution.startswith('3-Parameter Weibull'):
            low =  weibull_min.ppf(.00135, params['shape'], loc=params['threshold'], scale=params['scale'])
            med =  weibull_min.ppf(.5, params['shape'], loc=params['threshold'], scale=params['scale'])
            high =  weibull_min.ppf(.99865, params['shape'], loc=params['threshold'], scale=params['scale'])
        elif distribution.startswith('Gamma'):
            low = gamma.ppf(.00135, params['shape'], loc=params['threshold'], scale=params['scale'])
            med = gamma.ppf(.5, params['shape'], loc=params['threshold'], scale=params['scale'])
            high = gamma.ppf(.99865, params['shape'], loc=params['threshold'], scale=params['scale'])
        elif distribution.startswith('Logistic'):
            low = logistic.ppf(.00135, loc=params['location'], scale=params['scale'])
            med = logistic.ppf(.5, loc=params['location'], scale=params['scale'])
            high = logistic.ppf(.99865, loc=params['location'], scale=params['scale'])
        elif distribution.startswith('Largest Extreme Value'):
            low = gumbel_r.ppf(.00135, loc=params['location'], scale=params['scale'])
            med = gumbel_r.ppf(.5, loc=params['location'], scale=params['scale'])
            high = gumbel_r.ppf(.99865, loc=params['location'], scale=params['scale'])
        elif distribution.startswith('Smallest Extreme Value'):
            low = gumbel_l.ppf(.00135, loc=params['location'], scale=params['scale'])
            med = gumbel_l.ppf(.5, loc=params['location'], scale=params['scale'])
            high = gumbel_l.ppf(.99865, loc=params['location'], scale=params['scale'])
        elif distribution.startswith('Loglogistic'):
            low = fisk.ppf(.00135, c=1, loc=params['location'], scale=params['scale'])
            med= fisk.ppf(.5, c=1, loc=params['location'], scale=params['scale'])
            high = fisk.ppf(.99865, c=1, loc=params['location'], scale=params['scale'])
        elif distribution.startswith('3-Parameter Loglogistic'):
            low = fisk.ppf(.00135, c=1, loc=params['threshold'], scale=params['scale']) + params['location']
            med= fisk.ppf(.5, c=1, loc=params['threshold'], scale=params['scale']) + params['location']
            high = fisk.ppf(.99865, c=1, loc=params['threshold'], scale=params['scale']) + params['location']
        else:
            raise ValueError(f"Unsupported distribution: {distribution}")
        
        results[kpc] = {
            'dist': distribution,
            '.135th Percentile': low,
            '50th Percentile': med,
            '99.865th Percentile': high,
        }
    return results

def parse_tolerance(tolerance):
        range_pattern = re.compile(r'(?<!\S)(\d*\.\d+)\s*-\s*(\d*\.\d+)(?!\S)')
        specific_tolerance_pattern = re.compile(r'([A-Za-z ]+)\s+(\d*\.\d+)')
    
        range_match = range_pattern.search(tolerance)
    
        if range_match:
            max_val, min_val = map(float, range_match.groups())
            return ( max_val, min_val)
    
        specific_tolerance_match = specific_tolerance_pattern.search(tolerance)
        if specific_tolerance_match:
            tolerance_type, value = specific_tolerance_match.groups()
            return (0, float(value))
    
        return None, None

def info_dialog_init(parent, error, detail_text):
    dlg = QMessageBox(parent)
    dlg.setWindowTitle(error)
    dlg.setText(detail_text)
    dlg.setIcon(QMessageBox.Information)
    dlg.exec()
    
#_____________________________#
##Database helper functions##
#_____________________________#
def compare_features(existing_features, new_features): 
    """
    Compares existing features with new features and returns the updated list.
    Only the changes will be retained in the new feature list.

    Args:
        existing_features (dict): dictionary of existing features to compare
        new_features (dict): dictionary of new features 
    """
    updated_features = []
    
    for i, new_feature in enumerate(new_features):
        print(f"new feature: {new_feature}")
        if i < len(existing_features):
            updated_feature = {}
            for key, new_value in new_feature.items():
                if new_value and new_value != existing_features[i].get(key):
                    updated_feature[key] = new_value
            
            if updated_feature:
                updated_features.append({**existing_features[i], **updated_feature})
            else:
                updated_features.append(existing_features[i])
                
        else:
            updated_features.append(new_feature)
            
    return updated_features